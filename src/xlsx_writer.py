"""Generate a multi-tab XLSX from the list of parsed Entries."""
from __future__ import annotations

import os
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.models import Entry

# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')   # dark blue
_ALT_FILL      = PatternFill('solid', fgColor='DEEAF1')   # light blue
_SECTION_FILL  = PatternFill('solid', fgColor='2E75B6')   # medium blue
_TOTAL_FILL    = PatternFill('solid', fgColor='BDD7EE')   # pale blue
_POS_FILL      = PatternFill('solid', fgColor='C6EFCE')   # green
_NEG_FILL      = PatternFill('solid', fgColor='FFC7CE')   # red

_HEADER_FONT   = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
_TITLE_FONT    = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
_BOLD_FONT     = Font(bold=True, name='Calibri', size=10)
_NORMAL_FONT   = Font(name='Calibri', size=10)

_THIN = Side(style='thin', color='A0A0A0')
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_BRL_FMT = '#,##0.00'
_INT_FMT = '0'


# ── Main entry point ──────────────────────────────────────────────────────────

def write_xlsx(entries: list[Entry], output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    df = _entries_to_df(entries)

    _write_dados_brutos(wb, df)
    _write_resumo(wb, df)
    _write_totais(wb, df)
    _write_para_irpf(wb, df)

    wb.save(output_path)
    print(f'  Planilha salva em: {output_path}')


# ── DataFrame conversion ──────────────────────────────────────────────────────

_COLUMNS = [
    ('Arquivo',             'arquivo'),
    ('Instituição',         'instituicao'),
    ('CNPJ Instituição',    'cnpj_instituicao'),
    ('Ano-Calendário',      'ano_calendario'),
    ('Seção',               'secao'),
    ('Grupo',               'grupo'),
    ('Grupo Descrição',     'grupo_desc'),
    ('Código',              'codigo'),
    ('Código Descrição',    'codigo_desc'),
    ('Fonte Pagadora',      'fonte_pagadora'),
    ('CNPJ Fonte',          'cnpj_fonte'),
    ('Localização',         'localizacao'),
    ('Discriminação',       'discriminacao'),
    ('Valor 31/12/2024',    'valor_2024'),
    ('Valor 31/12/2025',    'valor_2025'),
    ('Rendimento',          'rendimento'),
    ('Tipo Rendimento',     'tipo_rendimento'),
    ('IRRF',                'irrf'),
    ('Observação',          'observacao'),
]

_COL_LABELS  = [c[0] for c in _COLUMNS]
_COL_ATTRS   = [c[1] for c in _COLUMNS]
_NUM_COLS    = {'Valor 31/12/2024', 'Valor 31/12/2025', 'Rendimento', 'IRRF',
                'Ano-Calendário'}


def _entries_to_df(entries: list[Entry]) -> pd.DataFrame:
    rows = [{label: getattr(e, attr) for label, attr in _COLUMNS} for e in entries]
    df = pd.DataFrame(rows, columns=_COL_LABELS)
    for col in ('Valor 31/12/2024', 'Valor 31/12/2025', 'Rendimento', 'IRRF'):
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    return df


# ── Helpers ───────────────────────────────────────────────────────────────────

def _style_header_row(ws, row_num: int, n_cols: int,
                      fill=_HEADER_FILL, font=_HEADER_FONT) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = _THIN_BORDER


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_df_rows(ws, df: pd.DataFrame, start_row: int,
                   num_col_names: set[str] | None = None,
                   alt_fill: PatternFill | None = _ALT_FILL) -> None:
    num_col_names = num_col_names or _NUM_COLS
    for row_idx, (_, row) in enumerate(df.iterrows()):
        fill = alt_fill if (row_idx % 2 == 1 and alt_fill) else None
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row + row_idx, column=col_idx, value=row[col_name])
            cell.font = _NORMAL_FONT
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            if col_name in num_col_names:
                cell.number_format = _BRL_FMT if col_name != 'Ano-Calendário' else _INT_FMT
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left', wrap_text=False)


def _add_total_row(ws, df: pd.DataFrame, start_row: int,
                   sum_cols: list[str], label: str = 'TOTAL') -> None:
    total_row = start_row + len(df)
    ws.cell(row=total_row, column=1, value=label)
    ws.cell(row=total_row, column=1).font = _BOLD_FONT
    ws.cell(row=total_row, column=1).fill = _TOTAL_FILL

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _BOLD_FONT
        cell.border = _THIN_BORDER
        if col_name in sum_cols:
            cell.value = df[col_name].sum()
            cell.number_format = _BRL_FMT
            cell.alignment = Alignment(horizontal='right')


# ── Tab 1: Dados Brutos ───────────────────────────────────────────────────────

def _write_dados_brutos(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet('Dados Brutos')
    ws.freeze_panes = 'A2'

    # Header
    for col_idx, label in enumerate(_COL_LABELS, 1):
        ws.cell(row=1, column=col_idx, value=label)
    _style_header_row(ws, 1, len(_COL_LABELS))
    ws.row_dimensions[1].height = 28

    _write_df_rows(ws, df, 2)

    widths = [26, 22, 18, 7, 24, 5, 28, 5, 46, 22, 18, 18, 34, 14, 14, 12, 20, 10, 30]
    _set_col_widths(ws, widths)

    # Excel Table for filtering
    tbl_ref = f'A1:{get_column_letter(len(_COL_LABELS))}{len(df) + 1}'
    tbl = Table(displayName='DadosBrutos', ref=tbl_ref)
    tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2',
                                        showFirstColumn=False,
                                        showLastColumn=False,
                                        showRowStripes=True)
    ws.add_table(tbl)


# ── Tab 2: Resumo ─────────────────────────────────────────────────────────────

def _write_resumo(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet('Resumo')

    # Pivot: rows = (Seção, Grupo, Código, Descrição), columns = Institution
    institutions = sorted(df['Instituição'].dropna().unique())

    pivot_rows: dict[tuple, dict] = {}
    for _, row in df.iterrows():
        key = (row['Seção'], row['Grupo'], row['Código Descrição'])
        if key not in pivot_rows:
            pivot_rows[key] = {inst: {'v24': 0.0, 'v25': 0.0, 'rend': 0.0}
                               for inst in institutions}
        pivot_rows[key][row['Instituição']]['v24'] += row['Valor 31/12/2024']
        pivot_rows[key][row['Instituição']]['v25'] += row['Valor 31/12/2025']
        pivot_rows[key][row['Instituição']]['rend'] += row['Rendimento']

    # Build header
    headers = ['Seção', 'Grupo/Código', 'Descrição']
    for inst in institutions:
        headers += [f'{inst} – 2024', f'{inst} – 2025', f'{inst} – Rendimento']
    headers += ['TOTAL 2024', 'TOTAL 2025', 'TOTAL Rendimentos']

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'D2'

    # Data rows
    current_section = None
    data_row = 2
    for row_idx, (key, inst_data) in enumerate(sorted(pivot_rows.items())):
        secao, grupo, desc = key
        if secao != current_section:
            current_section = secao
            # Section separator row
            ws.cell(row=data_row, column=1, value=secao.upper())
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=data_row, column=col)
                cell.fill = _SECTION_FILL
                cell.font = _TITLE_FONT
                cell.border = _THIN_BORDER
            data_row += 1

        fill = _ALT_FILL if row_idx % 2 == 1 else None

        ws.cell(row=data_row, column=1, value=secao).font = _NORMAL_FONT
        ws.cell(row=data_row, column=2, value=grupo).font = _NORMAL_FONT
        ws.cell(row=data_row, column=3, value=desc).font = _NORMAL_FONT

        total_v24 = total_v25 = total_rend = 0.0
        col = 4
        for inst in institutions:
            v24  = inst_data[inst]['v24']
            v25  = inst_data[inst]['v25']
            rend = inst_data[inst]['rend']
            total_v24  += v24
            total_v25  += v25
            total_rend += rend
            for val in (v24, v25, rend):
                cell = ws.cell(row=data_row, column=col, value=val if val else None)
                cell.number_format = _BRL_FMT
                cell.font = _NORMAL_FONT
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal='right')
                if fill:
                    cell.fill = fill
                col += 1

        for val in (total_v24, total_v25, total_rend):
            cell = ws.cell(row=data_row, column=col, value=val if val else None)
            cell.number_format = _BRL_FMT
            cell.font = _BOLD_FONT
            cell.border = _THIN_BORDER
            cell.fill = _TOTAL_FILL
            cell.alignment = Alignment(horizontal='right')
            col += 1

        for c in range(1, len(headers) + 1):
            ws.cell(row=data_row, column=c).border = _THIN_BORDER
            if fill and not ws.cell(row=data_row, column=c).fill.patternType == 'solid':
                ws.cell(row=data_row, column=c).fill = fill

        data_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 46
    for i in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


# ── Tab 3: Totais ─────────────────────────────────────────────────────────────

def _write_totais(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet('Totais')
    ws.freeze_panes = 'A2'

    agg = (
        df.groupby(['Seção', 'Grupo', 'Grupo Descrição', 'Código', 'Código Descrição'],
                   as_index=False)
        .agg({'Valor 31/12/2024': 'sum', 'Valor 31/12/2025': 'sum',
              'Rendimento': 'sum', 'IRRF': 'sum'})
        .sort_values(['Seção', 'Grupo', 'Código'])
    )

    headers = ['Seção', 'Grupo', 'Grupo Descrição', 'Código', 'Código Descrição',
               'Total 31/12/2024', 'Total 31/12/2025', 'Total Rendimentos', 'Total IRRF']
    num_h = {'Total 31/12/2024', 'Total 31/12/2025', 'Total Rendimentos', 'Total IRRF'}

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 28

    # Rename agg columns for _write_df_rows
    agg.columns = headers
    _write_df_rows(ws, agg, 2, num_col_names=num_h)

    _add_total_row(ws, agg, 2,
                   sum_cols=['Total 31/12/2024', 'Total 31/12/2025',
                              'Total Rendimentos', 'Total IRRF'])

    _set_col_widths(ws, [26, 5, 28, 5, 46, 16, 16, 16, 12])


# ── Tab 4: Para IRPF ─────────────────────────────────────────────────────────

def _write_para_irpf(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet('Para IRPF')

    row = 1
    # Title
    ws.cell(row=row, column=1, value='INFORMES DE RENDIMENTOS – RESUMO PARA IRPF')
    ws.merge_cells(f'A{row}:H{row}')
    title_cell = ws.cell(row=row, column=1)
    title_cell.font = Font(bold=True, size=13, name='Calibri', color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center')
    row += 2

    for inst in sorted(df['Instituição'].unique()):
        inst_df = df[df['Instituição'] == inst].copy()

        # Institution header
        cnpj_inst = inst_df['CNPJ Instituição'].iloc[0] if len(inst_df) else ''
        ws.cell(row=row, column=1, value=f'{inst}  (CNPJ: {cnpj_inst})')
        ws.merge_cells(f'A{row}:H{row}')
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = _SECTION_FILL
            cell.font = _TITLE_FONT
            cell.border = _THIN_BORDER
        row += 1

        # Column headers for this institution block
        sub_headers = ['Seção', 'Grupo', 'Código', 'Descrição',
                       'Valor 31/12/2024', 'Valor 31/12/2025', 'Rendimento', 'IRRF']
        for ci, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        row += 1

        for secao in sorted(inst_df['Seção'].unique()):
            sec_df = inst_df[inst_df['Seção'] == secao]

            for ri, (_, r) in enumerate(sec_df.iterrows()):
                fill = _ALT_FILL if ri % 2 == 1 else None
                vals = [
                    r['Seção'], r['Grupo'], r['Código'], r['Código Descrição'],
                    r['Valor 31/12/2024'], r['Valor 31/12/2025'],
                    r['Rendimento'], r['IRRF'],
                ]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = _NORMAL_FONT
                    cell.border = _THIN_BORDER
                    if fill:
                        cell.fill = fill
                    if ci >= 5:
                        cell.number_format = _BRL_FMT
                        cell.alignment = Alignment(horizontal='right')
                row += 1

            # Section subtotal
            ws.cell(row=row, column=4, value=f'Subtotal {secao}')
            ws.cell(row=row, column=4).font = _BOLD_FONT
            for ci, col in enumerate(['Valor 31/12/2024', 'Valor 31/12/2025',
                                       'Rendimento', 'IRRF'], 5):
                cell = ws.cell(row=row, column=ci, value=sec_df[col].sum())
                cell.font = _BOLD_FONT
                cell.fill = _TOTAL_FILL
                cell.number_format = _BRL_FMT
                cell.alignment = Alignment(horizontal='right')
                cell.border = _THIN_BORDER
            row += 1

        row += 1  # blank row between institutions

    # Column widths
    widths = [26, 6, 6, 46, 16, 16, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
