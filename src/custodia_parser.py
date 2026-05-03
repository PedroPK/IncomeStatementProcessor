"""
Parser for custom custódia data from XLSX file.

Handles user-provided custódia spreadsheet with:
- Ativo (ticker/asset code)
- Quantidade de Cotas
- Preço Médio (em 31/12/2025)

Generates Entry objects with calculated custo_aquisicao.
"""

from __future__ import annotations

import re
from pathlib import Path
import openpyxl
from src.models import Entry


def parse_custodia_xlsx(filepath: str, instituicao: str = 'Custódia Personalizada') -> list[Entry]:
    """
    Parse custom custódia data from XLSX file.
    
    Expected columns:
    - Column A: Ativo (ticker, e.g., PSSA3, PLAG11)
    - Column B: Quantidade de Cotas (integer or float)
    - Column C: Preço Médio (float, in BRL)
    
    Returns:
        List of Entry objects with calculated custo_aquisicao (quantidade × preço_médio)
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If required columns are missing or invalid data
    """
    
    if not Path(filepath).exists():
        raise FileNotFoundError(f"Custódia XLSX file not found: {filepath}")
    
    entries: list[Entry] = []
    
    # Load workbook and get active sheet
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
    except Exception as e:
        raise ValueError(f"Erro ao abrir XLSX: {e}")
    
    # Find header row (skip empty rows at top)
    header_row = None
    for row_idx in range(1, min(sheet.max_row + 1, 20)):
        cell_a = sheet.cell(row=row_idx, column=1).value
        if cell_a and isinstance(cell_a, str) and 'ativo' in cell_a.lower():
            header_row = row_idx
            break
    
    if header_row is None:
        # If no header found, assume data starts at row 2 (skip header in row 1)
        header_row = 1
        data_start_row = 2
    else:
        data_start_row = header_row + 1
    
    # Expected column positions (A=1, B=2, C=3)
    col_ativo = 1
    col_quantidade = 2
    col_preco_medio = 3
    
    # Parse data rows
    for row_idx in range(data_start_row, sheet.max_row + 1):
        ativo_cell = sheet.cell(row=row_idx, column=col_ativo).value
        qtd_cell = sheet.cell(row=row_idx, column=col_quantidade).value
        preco_cell = sheet.cell(row=row_idx, column=col_preco_medio).value
        
        # Stop if row is completely empty
        if not ativo_cell and not qtd_cell and not preco_cell:
            continue
        
        # Skip rows with missing data
        if not ativo_cell or not qtd_cell or not preco_cell:
            continue
        
        # Clean and validate data
        try:
            ativo = str(ativo_cell).strip().upper()
            quantidade = float(qtd_cell) if qtd_cell else 0.0
            preco_medio = float(preco_cell) if preco_cell else 0.0
        except (ValueError, TypeError) as e:
            print(f"  [aviso] Linha {row_idx} com dados inválidos: {e}")
            continue
        
        # Skip if quantity or price is zero
        if quantidade == 0 or preco_medio == 0:
            print(f"  [aviso] Linha {row_idx} ({ativo}): quantidade ou preço é zero")
            continue
        
        # Calculate custo de aquisição
        custo_aquisicao = quantidade * preco_medio
        
        # Map ticker to grupo/código (will be improved with fuzzy matching)
        grupo, codigo, grupo_desc, codigo_desc = _map_ativo_to_grupo_codigo(ativo)
        
        # Create Entry object
        entry = Entry(
            arquivo=Path(filepath).name,
            instituicao=instituicao,
            cnpj_instituicao='',
            ano_calendario=2025,
            secao='Bens e Direitos',
            grupo=grupo,
            grupo_desc=grupo_desc,
            codigo=codigo,
            codigo_desc=codigo_desc,
            fonte_pagadora=instituicao,
            cnpj_fonte='',
            localizacao='105 - Brasil',
            discriminacao=f'{ativo} – Ativo em Custódia',
            valor_2024=0.0,
            valor_2025=custo_aquisicao,  # Posição consolidada em 31/12/2025
            rendimento=0.0,  # Será preenchido por outro informe se houver
            tipo_rendimento='',
            irrf=0.0,
            observacao=f'Custódia: {quantidade:.2f} cotas × R$ {preco_medio:.2f}',
        )
        
        entries.append(entry)
        print(f"  ✅ {ativo}: {quantidade} cotas × R${preco_medio:.2f} = R${custo_aquisicao:,.2f}")
    
    return entries


def _map_ativo_to_grupo_codigo(ticker: str) -> tuple[str, str, str, str]:
    """
    Map ticker to IRPF grupo/código classification.
    
    Returns:
        Tuple of (grupo, codigo, grupo_desc, codigo_desc)
    """
    
    ticker_upper = ticker.upper().strip()
    
    # FII (Fundos Imobiliários) - terminam com 11
    if ticker_upper.endswith('11'):
        return '07', '02', 'Fundos de Investimento', 'Fundos Imobiliários'
    
    # Ações (terminam em números 3 ou 4)
    elif ticker_upper.endswith(('3', '4')):
        return '04', '01', 'Aplicações e Investimentos', 'Ações'
    
    # BDRs (terminam em 34 ou 35)
    elif ticker_upper.endswith(('34', '35')):
        return '03', '01', 'Ações de empresas', 'Ações de empresas no exterior'
    
    # ETFs (terminam em 11 e são fundos, mas podem ser outros - melhorar detecção)
    elif 'ETF' in ticker_upper or ticker_upper.endswith('11'):
        return '07', '03', 'Fundos de Investimento', 'ETFs (Exchange Traded Funds)'
    
    # Fundos em geral
    elif 'FUNDO' in ticker_upper or 'FDO' in ticker_upper:
        return '07', '01', 'Fundos de Investimento', 'Fundos de investimento'
    
    # Default: Aplicações e Investimentos (código 99 = outros)
    else:
        return '04', '99', 'Aplicações e Investimentos', f'Ativo: {ticker_upper}'
